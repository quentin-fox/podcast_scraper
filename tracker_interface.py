from blubrry import Blubrry
from openpyxl import load_workbook
import datetime
from pathlib import Path


def start_scraper():
    blubrry = Blubrry('rawtalkpodcast@gmail.com', 'k33p1tr4w', 0)
    blubrry.create_driver()
    blubrry.login()
    return(blubrry)


def get_data():
    track_sheet = tracker['Blubrry Automated Downloads']
    for row in track_sheet.iter_rows(min_row=2, max_col=6):
        if row[0].value and not row[4].value:
            ep_num = int(row[0].value)
            start = row[3].value
            end = start + datetime.timedelta(weeks=6)
            current_date = datetime.datetime.now()
            # row[4].value refers to total downloads
            if current_date >= end and not row[4].value:
                data = blubrry.scrape_episode_data(ep_num, start, end, 'Platforms')
                ep_data[ep_num] = data


def try_create_sheet(title, headers):
    try:
        sheet = tracker[title]
    except KeyError:
        tracker.create_sheet(title)
        sheet = tracker[title]
        sheet.append(headers)


def write_data():
    track_sheet = tracker['Blubrry Automated Downloads']
    geo_title = 'Geographic Tracking'
    dpc_title = 'Platform Tracking'
    try_create_sheet(geo_title, ['Episode #', 'Country', 'Downloads'])
    try_create_sheet(dpc_title, ['Episode #', 'Type', 'Downloads'])
    geo = tracker[geo_title]
    dpc = tracker[dpc_title]

    # writing main downloads data
    for row in track_sheet.iter_rows(min_row=2, max_col=11):
        ep_num = row[0].value
        if ep_num and not row[4].value:
            try:
                dl_data = ep_data[ep_num]['downloads'].values()
            except KeyError:
                continue
            val_cell = zip(dl_data, row[4:11])
            for val, cell in val_cell:
                cell.value = val

    # writing both geo and dpc data
    for ep_num, data in ep_data.items():
        # using list comprehensions for side effect
        geo_data = [[ep_num, country, num] for country, num in data['geo'].items()]
        [geo.append(row) for row in geo_data]
        dpc_data = [[ep_num, dpc, num] for dpc, num in data['dpc'].items()]
        [dpc.append(row) for row in dpc_data]


if __name__ == '__main__':
    tracker_path = Path('/Users/Quentin/iCloud/projects/podcast_scrape/Impact Reports/Raw Talk Podcast Analytics S4.xlsx')
    tracker = load_workbook(tracker_path.resolve(), data_only=True)
    blubrry = start_scraper()
    ep_data = {}
    get_data()
    write_data()
    tracker.save(tracker_path)
    blubrry.driver.quit()
    print('test output')
